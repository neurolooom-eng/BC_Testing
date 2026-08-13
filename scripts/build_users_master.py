"""
Builds data/users_master.xlsx — the local master roster of Bestcast users.

This workbook is the source of truth for WHO should have an account and
their profile details. It intentionally holds no passwords: Supabase Auth
owns credentials, this sheet owns identity/profile data and drives the
import script (scripts/import_users_from_excel.py).

Run: python scripts/build_users_master.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_PATH = "data/users_master.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="0D0B5C")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="F6F4EE")
EDITABLE_FILL = PatternFill("solid", fgColor="FFFFCC")
BODY_FONT = Font(name="Arial", size=10.5)
TITLE_FONT = Font(name="Arial", size=16, bold=True, color="0D0B5C")
NOTE_FONT = Font(name="Arial", size=10, italic=True, color="666666")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("User ID (login)", 20, "You choose this. Must match the Email column — the login page authenticates by email."),
    ("Full Name", 22, "First and last name."),
    ("Email", 26, "Real, working email address. Supabase sends the account-setup / password link here."),
    ("Role", 16, "e.g. Admin, Manager, Staff."),
    ("Department", 18, "e.g. Operations, Finance, Sales."),
    ("Status", 14, "Active, Inactive, or Pending. Only Active rows are invited by the import script."),
    ("Supabase User UID", 38, "Leave blank. Filled in automatically by the import script after the account is created."),
    ("Invite Sent Date", 16, "Leave blank. Filled in automatically by the import script."),
    ("Notes", 30, "Anything else worth tracking (start date, manager, etc.)."),
]

EXAMPLE_ROW = [
    "jsmith", "Jane Smith", "jane.smith@bestcast.com", "Manager",
    "Operations", "Active", "", "", "Example row — replace or delete.",
]


def build():
    wb = openpyxl.Workbook()

    # --- Legend sheet -----------------------------------------------------
    legend = wb.active
    legend.title = "Read Me First"
    legend["A1"] = "Bestcast — User Master Roster"
    legend["A1"].font = TITLE_FONT
    legend["A2"] = "How this workbook works"
    legend["A2"].font = Font(name="Arial", size=12, bold=True)

    instructions = [
        "1. Add one row per person on the 'Users' tab. Fill in the yellow columns only.",
        "2. 'User ID' and 'Email' must match — the sign-in page authenticates with the email address.",
        "3. Set Status to 'Active' for anyone who should be able to log in.",
        "4. Save the file, then run: python scripts/import_users_from_excel.py",
        "   This creates each Active user in Supabase Auth (no password stored here) and emails them",
        "   a secure link to set their own password. It also creates a matching row in the 'profiles'",
        "   table in Supabase (name, role, department) so the app knows who they are.",
        "5. The script writes the new Supabase User UID and the invite date back into this sheet — do",
        "   not hand-edit those two columns.",
        "6. To deactivate someone, set Status to 'Inactive' and re-run the script; it disables their",
        "   Supabase account instead of deleting their history.",
        "",
        "Security note: passwords are never stored in this spreadsheet or in Supabase in plain text.",
        "Supabase Auth stores only a salted hash, and this file never sees the password at all — users",
        "set it themselves via the emailed link.",
    ]
    row = 4
    for line in instructions:
        legend.cell(row=row, column=1, value=line).font = BODY_FONT if line else BODY_FONT
        row += 1
    legend.column_dimensions["A"].width = 100
    for r in range(4, row):
        legend.cell(row=r, column=1).alignment = Alignment(wrap_text=False)

    # --- Users sheet --------------------------------------------------------
    ws = wb.create_sheet("Users")
    for col_idx, (title, width, _note) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # notes row (row 2) describing each column
    for col_idx, (_title, _width, note) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=note)
        cell.font = NOTE_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = BORDER
    ws.row_dimensions[2].height = 45

    # example row
    for col_idx, value in enumerate(EXAMPLE_ROW, start=1):
        cell = ws.cell(row=3, column=col_idx, value=value)
        cell.font = Font(name="Arial", size=10.5, italic=True, color="888888")
        cell.border = BORDER
        if col_idx in (1, 2, 3, 4, 5, 6, 9):
            cell.fill = PatternFill("solid", fgColor="F6F4EE")

    # a handful of blank editable rows ready to fill in
    for r in range(4, 24):
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.font = BODY_FONT
            cell.border = BORDER
            if col_idx in (1, 2, 3, 4, 5, 6, 9):  # editable columns
                cell.fill = EDITABLE_FILL

    # data validation for Status column
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"Active,Inactive,Pending"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"F3:F200")

    ws.freeze_panes = "A3"

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    build()
