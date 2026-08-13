"""
Syncs data/users_master.xlsx (the "Users" tab) into Supabase:

  Status = Active    -> create the Supabase Auth user if missing (sends an
                         invite email so THEY set their own password — this
                         script never generates or stores a password), then
                         upsert a matching row in public.profiles.
  Status = Inactive   -> ban the existing Supabase Auth account so they can
                         no longer sign in, and mark profiles.status Inactive.
  Status = Pending    -> skipped. Pending rows are for self-service sign-ups
                         (public/signup.html) awaiting admin approval; flip
                         Status to Active once you're ready to admit them.

Writes the resulting Supabase User UID and invite date back into the sheet.

Requires two environment variables (never hardcode these):
  SUPABASE_URL              e.g. https://xxxx.supabase.co
  SUPABASE_SERVICE_ROLE_KEY the *service_role* key (Project Settings -> API).
                             This key bypasses Row Level Security -- keep it
                             out of git, out of the client, and out of any
                             file that isn't a local .env you've gitignored.

Usage:
  export SUPABASE_URL="https://xxxx.supabase.co"
  export SUPABASE_SERVICE_ROLE_KEY="ey..."
  python scripts/import_users_from_excel.py [path/to/users_master.xlsx]
"""
import os
import sys
from datetime import date

import openpyxl
import requests

XLSX_PATH = sys.argv[1] if len(sys.argv) > 1 else "data/users_master.xlsx"
SHEET_NAME = "Users"

SUPABASE_URL = os.environ.get("SUPABASE_URL", "").rstrip("/")
SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

COL = {
    "user_id": 1,
    "full_name": 2,
    "email": 3,
    "role": 4,
    "department": 5,
    "status": 6,
    "uid": 7,
    "invite_date": 8,
    "notes": 9,
}


def require_config():
    if not SUPABASE_URL or not SERVICE_KEY:
        sys.exit(
            "Missing SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY environment variables.\n"
            "See the header of this script for setup instructions."
        )


def admin_headers():
    return {
        "apikey": SERVICE_KEY,
        "Authorization": f"Bearer {SERVICE_KEY}",
        "Content-Type": "application/json",
    }


def find_user_by_email(email):
    resp = requests.get(
        f"{SUPABASE_URL}/auth/v1/admin/users",
        headers=admin_headers(),
        params={"email": email},
        timeout=30,
    )
    resp.raise_for_status()
    users = resp.json().get("users", [])
    return users[0] if users else None


def invite_user(email, full_name):
    resp = requests.post(
        f"{SUPABASE_URL}/auth/v1/invite",
        headers=admin_headers(),
        json={"email": email, "data": {"full_name": full_name}},
        timeout=30,
    )
    if resp.status_code >= 400:
        raise RuntimeError(f"invite failed for {email}: {resp.status_code} {resp.text}")
    return resp.json()


def set_ban(user_id, banned):
    resp = requests.put(
        f"{SUPABASE_URL}/auth/v1/admin/users/{user_id}",
        headers=admin_headers(),
        json={"ban_duration": "876000h" if banned else "none"},
        timeout=30,
    )
    resp.raise_for_status()


def upsert_profile(uid, user_id, full_name, email, role, department, status):
    resp = requests.post(
        f"{SUPABASE_URL}/rest/v1/profiles",
        headers={**admin_headers(), "Prefer": "resolution=merge-duplicates"},
        params={"on_conflict": "id"},
        json={
            "id": uid,
            "user_id": user_id,
            "full_name": full_name,
            "email": email,
            "role": role,
            "department": department,
            "status": status,
        },
        timeout=30,
    )
    if resp.status_code >= 400:
        raise RuntimeError(f"profile upsert failed for {email}: {resp.status_code} {resp.text}")


def main():
    require_config()

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[SHEET_NAME]

    changed = False
    for row in range(4, ws.max_row + 1):  # row 3 is the example row
        user_id = ws.cell(row=row, column=COL["user_id"]).value
        email = ws.cell(row=row, column=COL["email"]).value
        status = (ws.cell(row=row, column=COL["status"]).value or "").strip()
        if not email or not status:
            continue

        full_name = ws.cell(row=row, column=COL["full_name"]).value or ""
        role = ws.cell(row=row, column=COL["role"]).value
        department = ws.cell(row=row, column=COL["department"]).value
        existing_uid = ws.cell(row=row, column=COL["uid"]).value

        if status == "Pending":
            print(f"[skip] {email}: Pending, awaiting self-service sign-up/approval")
            continue

        if status == "Active":
            if existing_uid:
                uid = existing_uid
                set_ban(uid, banned=False)
                print(f"[reactivate] {email}")
            else:
                user = find_user_by_email(email)
                if user:
                    uid = user["id"]
                    print(f"[link] {email} already existed in Supabase Auth")
                else:
                    created = invite_user(email, full_name)
                    uid = created["id"]
                    print(f"[invite] {email}: invite email sent")
                ws.cell(row=row, column=COL["uid"], value=uid)
                ws.cell(row=row, column=COL["invite_date"], value=date.today().isoformat())
                changed = True

            upsert_profile(uid, user_id, full_name, email, role, department, "Active")

        elif status == "Inactive":
            if existing_uid:
                set_ban(existing_uid, banned=True)
                upsert_profile(existing_uid, user_id, full_name, email, role, department, "Inactive")
                print(f"[disable] {email}")
            else:
                print(f"[skip] {email}: Inactive with no Supabase account yet, nothing to disable")

        else:
            print(f"[warn] {email}: unrecognized Status '{status}', skipping")

    if changed:
        wb.save(XLSX_PATH)
        print(f"Wrote UIDs / invite dates back to {XLSX_PATH}")

    print("Done.")


if __name__ == "__main__":
    main()
